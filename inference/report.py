"""Payload and Excel report generation for rice grading results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .grader import GRADE_THRESHOLDS, PARAMETER_ORDER


def _iso_now() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def build_payload(
    grader_result: dict[str, Any],
    enriched_grains: list[dict[str, Any]],
    stub_mode: bool,
) -> dict[str, Any]:
    return {
        "timestamp": _iso_now(),
        "grade": grader_result["grade"],
        "limiting_factor": grader_result["limiting_factor"],
        "grain_size_class": grader_result["grain_size_class"],
        "total_grains_detected": int(grader_result["total_grains"]),
        "stub_mode": bool(stub_mode),
        "parameters": grader_result["parameters"],
        "per_grain": [
            {
                "grain_id": int(g["grain_id"]),
                "class_label": str(g["class_label"]),
                "length_mm": float(g["length_mm"]),
                "width_mm": float(g["width_mm"]),
                "area_px": float(g["area_px"]),
                "grain_size_class": str(g["grain_size_class"]),
                "ir_mean_intensity": float(g["ir_mean_intensity"]),
            }
            for g in sorted(enriched_grains, key=lambda item: int(item["grain_id"]))
        ],
    }


def _grade_threshold_for_display(grade: str) -> dict[str, float]:
    if grade in GRADE_THRESHOLDS:
        return GRADE_THRESHOLDS[grade]
    return GRADE_THRESHOLDS["Grade no. 5"]


def _autofit_columns(sheet) -> None:
    for col in sheet.iter_cols():
        max_len = 0
        first_cell = next((cell for cell in col if hasattr(cell, "column_letter")), None)
        if first_cell is None:
            continue
        col_letter = first_cell.column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        sheet.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)


def save_excel(payload: dict[str, Any], grader_result: dict[str, Any]) -> str:
    out_dir = Path("outputs")
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"report_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Rice Grading Report - PNS/BAFS 290:2025"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"], ws["B2"] = "Timestamp:", payload["timestamp"]
    ws["A3"], ws["B3"] = "Grade:", payload["grade"]
    ws["A4"], ws["B4"] = "Limiting Factor:", payload["limiting_factor"]
    ws["A5"], ws["B5"] = "Grain Size Class:", payload["grain_size_class"]
    ws["A6"], ws["B6"] = "Total Grains:", payload["total_grains_detected"]
    ws["A7"], ws["B7"] = "Mode:", "STUB (synthetic data)" if payload["stub_mode"] else "REAL MODEL"

    grade = payload["grade"]
    if grade in {"Premium", "Grade no. 1"}:
        ws["B3"].font = Font(bold=True, color="008000")
    elif grade in {"Grade no. 2", "Grade no. 3"}:
        ws["B3"].font = Font(bold=True, color="B8860B")
    else:
        ws["B3"].font = Font(bold=True, color="B22222")

    ws["A9"], ws["B9"], ws["C9"], ws["D9"], ws["E9"] = (
        "Parameter",
        "Actual (%)",
        "Threshold (%)",
        "Margin (%)",
        "Status",
    )
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}9"].font = Font(bold=True)

    thresholds = _grade_threshold_for_display(payload["grade"])
    params = grader_result["parameters"]
    rows: list[tuple[str, float, float, float, str]] = []
    for param in PARAMETER_ORDER:
        actual = float(params[param])
        threshold = float(thresholds[param])
        margin = threshold - actual
        status = "PASS" if actual <= threshold else "FAIL"
        rows.append((param, actual, threshold, margin, status))

    rows.sort(key=lambda item: item[3])
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    limiting_fill = PatternFill("solid", fgColor="FFF2CC")

    for idx, (param, actual, threshold, margin, status) in enumerate(rows, start=10):
        ws[f"A{idx}"] = param
        ws[f"B{idx}"] = round(actual, 4)
        ws[f"C{idx}"] = round(threshold, 4)
        ws[f"D{idx}"] = round(margin, 4)
        ws[f"E{idx}"] = status
        ws[f"E{idx}"].fill = pass_fill if status == "PASS" else fail_fill
        if param == payload["limiting_factor"] and status == "PASS":
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{idx}"].fill = limiting_fill

    ws2 = wb.create_sheet("Per-Grain Detail")
    headers = [
        "Grain ID",
        "Class Label",
        "Length (mm)",
        "Width (mm)",
        "Area (px^2)",
        "Size Class",
        "IR Intensity",
    ]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)

    for grain in payload["per_grain"]:
        ws2.append(
            [
                grain["grain_id"],
                grain["class_label"],
                round(grain["length_mm"], 4),
                round(grain["width_mm"], 4),
                round(grain["area_px"], 2),
                grain["grain_size_class"],
                round(grain["ir_mean_intensity"], 2),
            ]
        )
    ws2.freeze_panes = "A2"

    _autofit_columns(ws)
    _autofit_columns(ws2)
    wb.save(out_path)
    return str(out_path)


def build_report(
    grader_result: dict[str, Any],
    enriched_grains: list[dict[str, Any]],
    stub_mode: bool,
) -> tuple[dict[str, Any], str]:
    payload = build_payload(grader_result, enriched_grains, stub_mode)
    excel_path = save_excel(payload, grader_result)
    return payload, excel_path
